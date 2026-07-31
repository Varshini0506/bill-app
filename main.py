from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
import os
import threading
import sys
import webbrowser
import io
import time
import json

# =============================
# Resource Path (for EXE build)
# =============================
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

# =============================
# Flask Setup
# =============================
app = Flask(
    __name__,
    template_folder=resource_path("templates"),
    static_folder=resource_path("static")
)

# =============================
# Storage Folder
# =============================
FOLDER_PATH = r"C:\SRI Traders"
os.makedirs(FOLDER_PATH, exist_ok=True)

FILE_PATH = os.path.join(FOLDER_PATH, "invoices.xlsx")
INVENTORY_FILE_PATH = os.path.join(FOLDER_PATH, "inventory.xlsx")

# =============================
# Invoice Excel Functions
# =============================

def ensure_excel():
    if not os.path.exists(FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["Invoice ID", "File Name", "HTML Data", "Bill Items JSON", "Status"])
        wb.save(FILE_PATH)


def load_invoices():
    ensure_excel()
    invoices = []
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        invoice_id = row[0]
        file_name = row[1] if len(row) > 1 else ""
        html_data = row[2] if len(row) > 2 else ""
        bill_items = row[3] if len(row) > 3 and row[3] else ""
        status = row[4] if len(row) > 4 and row[4] else "Active"
        if status in ("Cancelled", "Deleted"):
            continue
        invoices.append({
            "id": invoice_id,
            "name": file_name,
            "data": html_data,
            "bill_items": bill_items,
            "status": status
        })
    return invoices


def save_invoice(file_name, html_data, bill_items_json="", status="Active"):
    ensure_excel()
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    invoice_id = str(int(time.time() * 1000))  # unique timestamp
    ws.append([invoice_id, file_name, html_data, bill_items_json, status])
    wb.save(FILE_PATH)
    return invoice_id


def get_invoice_bill_items(invoice_id):
    """Return the Bill Items JSON string for the given invoice_id."""
    ensure_excel()
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None and str(row[0]) == str(invoice_id):
            return row[3] if len(row) > 3 and row[3] else ""
    return ""


def update_invoice_excel(invoice_id, file_name, html_data, bill_items_json="", status="Active"):
    ensure_excel()
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(invoice_id):
            row[1].value = file_name
            row[2].value = html_data
            if len(row) > 3:
                row[3].value = bill_items_json
            if len(row) > 4:
                row[4].value = status
            else:
                ws.cell(row=row[0].row, column=5, value=status)
            break
    wb.save(FILE_PATH)


# =============================
# Inventory Excel Functions & Validation
# =============================

def ensure_inventory_excel():
    if not os.path.exists(INVENTORY_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["Product ID", "Product Name", "Purchase Quantity", "Sold Quantity", "Purchase Date", "Unit Price", "Notes", "GST", "Cost Price", "Monthly Sold JSON"])
        wb.save(INVENTORY_FILE_PATH)


def load_products():
    ensure_inventory_excel()
    products = []
    wb = load_workbook(INVENTORY_FILE_PATH)
    ws = wb.active
    current_month = time.strftime("%Y-%m")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        product_id = row[0]
        name = row[1] if len(row) > 1 else ""
        purchase_qty = row[2] if len(row) > 2 else 0
        sold_qty = row[3] if len(row) > 3 else 0
        purchase_date = row[4] if len(row) > 4 else ""
        unit_price = row[5] if len(row) > 5 else None
        notes = row[6] if len(row) > 6 else ""
        gst = row[7] if len(row) > 7 else "5%"
        cost_price = row[8] if len(row) > 8 and row[8] is not None else 0.0
        monthly_sold_json = row[9] if len(row) > 9 and row[9] is not None else "{}"

        monthly_sold_qty = 0
        try:
            m_dict = json.loads(str(monthly_sold_json))
            monthly_sold_qty = int(m_dict.get(current_month, 0))
        except Exception:
            monthly_sold_qty = 0

        products.append({
            "id": str(product_id),
            "name": str(name),
            "purchase_qty": int(purchase_qty) if purchase_qty is not None else 0,
            "sold_qty": int(sold_qty) if sold_qty is not None else 0,
            "purchase_date": str(purchase_date) if purchase_date is not None else "",
            "unit_price": float(unit_price) if unit_price is not None else None,
            "notes": str(notes) if notes is not None else "",
            "gst": str(gst) if gst is not None else "5%",
            "cost_price": float(cost_price) if cost_price is not None else 0.0,
            "monthly_sold": monthly_sold_qty
        })
    return products


def save_product_excel(name, purchase_qty, sold_qty, purchase_date, unit_price, notes, cost_price=0.0):
    ensure_inventory_excel()
    wb = load_workbook(INVENTORY_FILE_PATH)
    ws = wb.active
    product_id = str(int(time.time() * 1000))
    ws.append([product_id, name, purchase_qty, sold_qty, purchase_date, unit_price, notes, "5%", cost_price, "{}"])
    wb.save(INVENTORY_FILE_PATH)
    return product_id


def update_product_excel(product_id, name, purchase_qty, sold_qty, purchase_date, unit_price, notes, cost_price=0.0):
    ensure_inventory_excel()
    wb = load_workbook(INVENTORY_FILE_PATH)
    ws = wb.active
    found = False
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(product_id):
            row[1].value = name
            row[2].value = purchase_qty
            row[3].value = sold_qty
            row[4].value = purchase_date
            row[5].value = unit_price
            row[6].value = notes
            row[7].value = "5%"
            
            # Ensure columns 9 (Cost Price) and 10 (Monthly Sold JSON) exist
            if len(row) > 8:
                row[8].value = cost_price
            else:
                ws.cell(row=row[0].row, column=9, value=cost_price)

            found = True
            break
    wb.save(INVENTORY_FILE_PATH)
    return found


def delete_product_excel(product_id):
    ensure_inventory_excel()
    wb = load_workbook(INVENTORY_FILE_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    ws.delete_rows(1, ws.max_row)
    ws.append(["Product ID", "Product Name", "Purchase Quantity", "Sold Quantity", "Purchase Date", "Unit Price", "Notes", "GST", "Cost Price", "Monthly Sold JSON"])
    deleted = False
    for r in rows[1:]:
        if r[0] is not None:
            if str(r[0]) != str(product_id):
                ws.append(r)
            else:
                deleted = True
    wb.save(INVENTORY_FILE_PATH)
    return deleted


def validate_product_data(data):
    name = data.get("name")
    purchase_qty = data.get("purchase_qty")
    sold_qty = data.get("sold_qty", 0)

    if not name or str(name).strip() == "":
        return False, "Product Name is required"

    try:
        purchase_qty = int(purchase_qty)
        if purchase_qty < 0:
            return False, "Purchase Quantity cannot be negative"
    except (ValueError, TypeError):
        return False, "Invalid Purchase Quantity"

    try:
        sold_qty = int(sold_qty)
        if sold_qty < 0:
            return False, "Sold Quantity cannot be negative"
    except (ValueError, TypeError):
        return False, "Invalid Sold Quantity"

    if sold_qty > purchase_qty:
        return False, "Sold Quantity cannot be greater than Purchased Quantity"

    unit_price = data.get("unit_price")
    if unit_price is not None and unit_price != "":
        try:
            unit_price = float(unit_price)
            if unit_price < 0:
                return False, "Selling Price (Unit Price) cannot be negative"
        except ValueError:
            return False, "Invalid Selling Price"

    cost_price = data.get("cost_price")
    if cost_price is not None and cost_price != "":
        try:
            cost_price = float(cost_price)
            if cost_price < 0:
                return False, "Cost Price cannot be negative"
        except ValueError:
            return False, "Invalid Cost Price"

    return True, ""


# =============================
# Inventory Integration Helpers
# =============================

def update_sold_qty_by_name(product_name, delta):
    """
    Update sold_qty and monthly_sold for a product matched by name (case-insensitive).
    delta > 0 : increase sold qty (bill generated — stock decreases)
    delta < 0 : decrease sold qty (bill deleted/reverted — stock restored)
    """
    ensure_inventory_excel()
    wb = load_workbook(INVENTORY_FILE_PATH)
    ws = wb.active
    current_month = time.strftime("%Y-%m")
    for row in ws.iter_rows(min_row=2):
        if row[1].value and str(row[1].value).strip().lower() == str(product_name).strip().lower():
            current_sold = int(row[3].value) if row[3].value is not None else 0
            purchase_qty = int(row[2].value) if row[2].value is not None else 0
            new_sold = current_sold + delta
            new_sold = max(0, new_sold)           # never go below 0
            new_sold = min(new_sold, purchase_qty) # never exceed purchased
            row[3].value = new_sold

            # Update monthly sold JSON
            monthly_json_str = str(row[9].value) if len(row) > 9 and row[9].value is not None else "{}"
            monthly_dict = {}
            try:
                monthly_dict = json.loads(monthly_json_str)
            except Exception:
                monthly_dict = {}
            
            cur_month_val = int(monthly_dict.get(current_month, 0))
            new_month_val = max(0, cur_month_val + delta)
            monthly_dict[current_month] = new_month_val

            if len(row) > 9:
                row[9].value = json.dumps(monthly_dict)
            else:
                ws.cell(row=row[0].row, column=10, value=json.dumps(monthly_dict))

            wb.save(INVENTORY_FILE_PATH)
            return True
    return False


def deduct_inventory(bill_items_json):
    """Deduct sold quantities from inventory based on bill items JSON string."""
    if not bill_items_json:
        return
    try:
        items = json.loads(bill_items_json)
        for item in items:
            name = item.get("name", "")
            qty = int(float(item.get("qty", 0)))
            if name and qty > 0:
                update_sold_qty_by_name(name, qty)
    except Exception:
        pass


def revert_inventory(bill_items_json):
    """Restore sold quantities to inventory (reverse of deduct_inventory)."""
    if not bill_items_json:
        return
    try:
        items = json.loads(bill_items_json)
        for item in items:
            name = item.get("name", "")
            qty = int(float(item.get("qty", 0)))
            if name and qty > 0:
                update_sold_qty_by_name(name, -qty)
    except Exception:
        pass


# =============================
# Routes
# =============================

@app.route("/")
def index():
    return render_template("index.html")


# Get all invoices (sidebar)
@app.route("/get-invoices")
def get_invoices():
    invoices = load_invoices()
    return jsonify(invoices)


# Get one invoice
@app.route("/get-invoice/<invoice_id>")
def get_invoice(invoice_id):
    ensure_excel()
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(invoice_id):
            return jsonify({
                "id": row[0],
                "name": row[1],
                "data": row[2],
                "bill_items": row[3] if len(row) > 3 and row[3] else "",
                "status": row[4] if len(row) > 4 and row[4] else "Active"
            })
    return jsonify({"data": ""})


# Cancel invoice
@app.route("/cancel-invoice", methods=["POST"])
def cancel_invoice():
    data = request.json
    cancel_id = data.get("id")

    # Revert inventory before deleting the bill
    old_items_json = get_invoice_bill_items(cancel_id)
    if old_items_json:
        revert_inventory(old_items_json)

    wb = load_workbook(FILE_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    ws.delete_rows(1, ws.max_row)
    ws.append(["Invoice ID", "File Name", "HTML Data", "Bill Items JSON", "Status"])

    for r in rows[1:]:
        if r[0] is not None:
            if str(r[0]) != str(cancel_id):
                row_list = list(r)
                while len(row_list) < 5:
                    row_list.append("")
                ws.append(row_list)

    wb.save(FILE_PATH)
    return jsonify({"message": "Invoice cancelled"})


# Save new invoice
@app.route("/save", methods=["POST"])
def save():
    data = request.json
    file_name = data.get("name")
    html_data = data.get("data")
    bill_items = data.get("bill_items", [])
    bill_items_json = json.dumps(bill_items)

    invoice_id = save_invoice(file_name, html_data, bill_items_json)
    deduct_inventory(bill_items_json)  # auto-deduct stock

    return jsonify({
        "message": "Invoice saved",
        "invoice_id": invoice_id
    })


# Update invoice
@app.route("/update-invoice", methods=["POST"])
def update_invoice():
    data = request.json
    invoice_id = data.get("id")
    file_name = data.get("name")
    html_data = data.get("data")
    bill_items = data.get("bill_items", [])
    new_bill_items_json = json.dumps(bill_items)

    # Revert old inventory quantities first
    old_items_json = get_invoice_bill_items(invoice_id)
    if old_items_json:
        revert_inventory(old_items_json)

    # Save updated invoice
    update_invoice_excel(invoice_id, file_name, html_data, new_bill_items_json)

    # Apply new inventory deductions
    deduct_inventory(new_bill_items_json)

    return jsonify({"message": "Invoice updated"})


# PDF relay (used for printing)
@app.route("/pdf-action", methods=["POST"])
def pdf_action():
    file = request.files["file"]
    pdf_bytes = file.read()
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=False,
        download_name="invoice.pdf"
    )


@app.route("/rename-invoice", methods=["POST"])
def rename_invoice():
    data = request.json
    old_id = data["old_id"]
    new_id = data["new_id"]

    wb = load_workbook(FILE_PATH)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == old_id:
            row[0].value = new_id
            break

    wb.save(FILE_PATH)
    return jsonify({"message": "Renamed"})


@app.route("/delete-invoice", methods=["POST"])
def delete_invoice():
    data = request.json
    delete_id = data["id"]

    # Revert inventory before deleting the bill
    old_items_json = get_invoice_bill_items(delete_id)
    if old_items_json:
        revert_inventory(old_items_json)

    wb = load_workbook(FILE_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    ws.delete_rows(1, ws.max_row)
    ws.append(["Invoice ID", "File Name", "HTML Data", "Bill Items JSON", "Status"])

    for r in rows[1:]:
        if r[0] is not None:
            if str(r[0]) != str(delete_id):
                row_list = list(r)
                while len(row_list) < 5:
                    row_list.append("")
                ws.append(row_list)

    wb.save(FILE_PATH)
    return jsonify({"message": "Deleted"})


# =============================
# Inventory Routes
# =============================

@app.route("/inventory")
def inventory_page():
    return render_template("inventory.html")


@app.route("/get-products")
def get_products():
    return jsonify(load_products())


@app.route("/get-product-names")
def get_product_names():
    """Simplified product list for billing dropdown and stock validation."""
    products = load_products()
    result = []
    for p in products:
        available = p["purchase_qty"] - p["sold_qty"]
        result.append({
            "id": p["id"],
            "name": p["name"],
            "available_qty": available,
            "unit_price": p["unit_price"],
            "gst": p["gst"]
        })
    return jsonify(result)


@app.route("/save-product", methods=["POST"])
def save_product_route():
    data = request.json
    is_valid, msg = validate_product_data(data)
    if not is_valid:
        return jsonify({"success": False, "message": msg}), 400

    name = str(data.get("name")).strip()
    purchase_qty = int(data.get("purchase_qty"))
    sold_qty = int(data.get("sold_qty", 0))
    purchase_date = str(data.get("purchase_date", ""))
    unit_price = data.get("unit_price")
    if unit_price is not None and unit_price != "":
        unit_price = float(unit_price)
    else:
        unit_price = None
    cost_price = data.get("cost_price")
    if cost_price is not None and cost_price != "":
        cost_price = float(cost_price)
    else:
        cost_price = 0.0
    notes = str(data.get("notes", ""))

    product_id = save_product_excel(name, purchase_qty, sold_qty, purchase_date, unit_price, notes, cost_price)
    return jsonify({"success": True, "message": "Product saved successfully", "product_id": product_id})


@app.route("/update-product", methods=["POST"])
def update_product_route():
    data = request.json
    is_valid, msg = validate_product_data(data)
    if not is_valid:
        return jsonify({"success": False, "message": msg}), 400

    product_id = data.get("id")
    if not product_id:
        return jsonify({"success": False, "message": "Product ID is required"}), 400

    name = str(data.get("name")).strip()
    purchase_qty = int(data.get("purchase_qty"))
    sold_qty = int(data.get("sold_qty", 0))
    purchase_date = str(data.get("purchase_date", ""))
    unit_price = data.get("unit_price")
    if unit_price is not None and unit_price != "":
        unit_price = float(unit_price)
    else:
        unit_price = None
    cost_price = data.get("cost_price")
    if cost_price is not None and cost_price != "":
        cost_price = float(cost_price)
    else:
        cost_price = 0.0
    notes = str(data.get("notes", ""))

    found = update_product_excel(product_id, name, purchase_qty, sold_qty, purchase_date, unit_price, notes, cost_price)
    if found:
        return jsonify({"success": True, "message": "Product updated successfully"})
    else:
        return jsonify({"success": False, "message": "Product not found"}), 404


@app.route("/delete-product", methods=["POST"])
def delete_product_route():
    data = request.json
    product_id = data.get("id")
    if not product_id:
        return jsonify({"success": False, "message": "Product ID is required"}), 400

    deleted = delete_product_excel(product_id)
    if deleted:
        return jsonify({"success": True, "message": "Product deleted successfully"})
    else:
        return jsonify({"success": False, "message": "Product not found"}), 404


# =============================
# Start Flask Server
# =============================

def start_flask():
    app.run(
        host="127.0.0.1",
        port=5000,
        debug=False
    )

# =============================
# Auto Open Browser
# =============================
def open_browser():
    webbrowser.open("http://127.0.0.1:5000/")

if __name__ == "__main__":
    threading.Timer(1.5, open_browser).start()
    app.run(host="127.0.0.1", port=5000, debug=False)


# py -3.12 -m PyInstaller --onefile --noconsole --add-data "templates;templates" --add-data "static;static" main.py